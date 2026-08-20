import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from aiogram.types import BufferedInputFile

def generate_excel_report(registrations: list) -> BufferedInputFile:
    """Barcha ro'yxatdan o'tganlar ma'lumotlarini Excel fayliga o'tkazadi va BufferedInputFile obyektini qaytaradi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Tashrif buyuruvchilar"

    # Sarlavhalar
    headers = [
        "ID", "Telegram ID", "Username", "Ism va Familiya", "Lavozim",
        "Kompaniya nomi", "Faoliyat sohasi", "Mamlakat", "Shahar",
        "Telefon raqami", "Email", "Kelish kuni", "Tashrif maqsadi",
        "Qo'shimcha izoh", "Ro'yxatdan o mevaqti"
    ]

    # Sarlavha stili
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.append(headers)

    # Sarlavhalarni formatlash
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Ma'lumotlarni yozish
    for row_idx, item in enumerate(registrations, 2):
        row_data = [
            item.get("id"),
            item.get("telegram_id"),
            f"@{item.get('username')}" if item.get("username") else "",
            item.get("full_name"),
            item.get("position"),
            item.get("company"),
            item.get("industry"),
            item.get("country"),
            item.get("city"),
            item.get("phone"),
            item.get("email"),
            item.get("visit_day"),
            item.get("visit_purpose"),
            item.get("comment"),
            str(item.get("created_at"))
        ]
        ws.append(row_data)

        # Kataklarni formatlash
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 2, 8, 9, 10, 12, 15]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Ustun kengliklarini avtomatik moslashtirish
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Xotiraga saqlash
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return BufferedInputFile(file_stream.getvalue(), filename="SOF_EXPO_B2B_Registrations.xlsx")
